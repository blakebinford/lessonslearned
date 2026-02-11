"""
Parsers for importing lessons learned from XLSX, XLS, and CSV files.
Supports the Strike F428-style format and generic spreadsheet layouts.
"""
import io
import csv
import openpyxl


DISCIPLINE_MAP = {
    "weld": "Welding",
    "coat": "Coatings",
    "nde": "NDE",
    "ndt": "NDE",
    "inspect": "NDE",
    "environment": "Environmental",
    "safety": "Safety",
    "civil": "Civil",
    "mechanical": "Mechanical",
    "electric": "Electrical",
    "procur": "Materials/Procurement",
    "project": "Project Controls",
    "manage": "Project Controls",
    "survey": "Civil",
    "quality": "Quality",
    "regulat": "Regulatory",
}

SEVERITY_KEYWORDS = {
    "Critical": ["critical", "safety", "injury", "fatality", "catastroph"],
    "High": ["week", "month", "significant", "$", "major", "substantial"],
    "Medium": ["delay", "rework", "slow", "moderate", "minor cost"],
}


def map_discipline(raw):
    """Map a raw discipline/category string to a standard discipline."""
    if not raw:
        return ""
    lower = raw.lower()
    for keyword, discipline in DISCIPLINE_MAP.items():
        if keyword in lower:
            return discipline
    return ""


def map_severity(impact_text):
    """Infer severity from impact description text."""
    if not impact_text:
        return "Medium"
    lower = impact_text.lower()
    for severity, keywords in SEVERITY_KEYWORDS.items():
        if any(kw in lower for kw in keywords):
            return severity
    return "Medium"


def find_col(headers, *terms):
    """Find column index by fuzzy matching header names."""
    for term in terms:
        for i, h in enumerate(headers):
            if term in h:
                return i
    return -1


def cell_str(value):
    """Convert cell value to clean string."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("none", "nan", "nat"):
        return ""
    return s


def parse_xlsx(file_bytes):
    """
    Parse an XLSX file and extract lessons learned.
    Returns a list of dicts ready for Lesson model creation.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    # Find the lessons sheet
    sheet = None
    for name in wb.sheetnames:
        lower = name.lower()
        if "lesson" in lower or "experience" in lower:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], f'Sheet "{sheet.title}" has no data rows.'

    # Map headers
    raw_headers = [cell_str(h).lower() for h in rows[0]]

    col_map = {
        "id": find_col(raw_headers, "ll #", "ll#", "id", "number"),
        "date": find_col(raw_headers, "date logged", "date"),
        "project": find_col(raw_headers, "project name", "project"),
        "region": find_col(raw_headers, "region", "location"),
        "discipline": find_col(raw_headers, "discipline"),
        "category": find_col(raw_headers, "category"),
        "phase": find_col(raw_headers, "phase", "milestone"),
        "logged_by": find_col(raw_headers, "logged by", "author"),
        "context": find_col(raw_headers, "situation", "context"),
        "what_happened": find_col(raw_headers, "what happened", "event", "description"),
        "impact": find_col(raw_headers, "impact", "consequence"),
        "root_cause": find_col(raw_headers, "root cause"),
        "worked_didnt": find_col(raw_headers, "what worked", "worked/didn"),
        "recommendation": find_col(raw_headers, "recommendation", "action"),
        "keywords": find_col(raw_headers, "keyword", "tag"),
        "status": find_col(raw_headers, "status"),
        "assigned_to": find_col(raw_headers, "assigned to"),
        "docs": find_col(raw_headers, "supporting doc", "document"),
    }

    def get(row, key):
        idx = col_map.get(key, -1)
        if idx < 0 or idx >= len(row):
            return ""
        return cell_str(row[idx])

    lessons = []
    for i, row in enumerate(rows[1:], start=2):
        context = get(row, "context")
        what_happened = get(row, "what_happened")
        if not context and not what_happened:
            continue

        # Build description
        description = context
        if what_happened and what_happened != context:
            description += f"\n\nWhat happened: {what_happened}" if description else what_happened
        worked = get(row, "worked_didnt")
        if worked:
            description += f"\n\n{worked}"

        # Build title from first sentence of context
        title = context.split(".")[0].split("!")[0].split("\n")[0].strip()
        if len(title) > 100:
            title = title[:97] + "..."
        if not title:
            title = what_happened.split(".")[0].split("\n")[0].strip()[:100]
        if not title:
            continue

        discipline_raw = get(row, "discipline") or get(row, "category")
        impact = get(row, "impact")
        logged_by = get(row, "logged_by")
        status = get(row, "status")
        assigned_to = get(row, "assigned_to")
        docs = get(row, "docs")
        keywords_parts = [get(row, "keywords"), get(row, "category")]
        if logged_by:
            keywords_parts.append(f"logged by: {logged_by}")
        if status:
            keywords_parts.append(f"status: {status}")
        if docs:
            keywords_parts.append(docs)

        lessons.append({
            "title": title,
            "description": description,
            "root_cause": get(row, "root_cause"),
            "recommendation": get(row, "recommendation"),
            "impact": impact,
            "work_type": "Pipeline Construction",
            "phase": get(row, "phase"),
            "discipline": map_discipline(discipline_raw),
            "severity": map_severity(impact),
            "environment": "",
            "project": get(row, "project"),
            "location": get(row, "region"),
            "keywords": ", ".join(filter(None, keywords_parts)),
            "logged_by": logged_by,
            "status": status,
            "assigned_to": assigned_to,
            "supporting_docs": docs,
        })

    return lessons, None


def parse_csv(file_bytes):
    """Parse a CSV file and extract lessons. Same logic as XLSX but CSV source."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if len(rows) < 2:
        return [], "CSV has no data rows."

    # Reuse XLSX logic by converting to same format
    raw_headers = [h.lower().strip() for h in rows[0]]

    col_map = {
        "id": find_col(raw_headers, "ll #", "ll#", "id", "number"),
        "date": find_col(raw_headers, "date logged", "date"),
        "project": find_col(raw_headers, "project name", "project"),
        "region": find_col(raw_headers, "region", "location"),
        "discipline": find_col(raw_headers, "discipline"),
        "category": find_col(raw_headers, "category"),
        "phase": find_col(raw_headers, "phase", "milestone"),
        "logged_by": find_col(raw_headers, "logged by", "author"),
        "context": find_col(raw_headers, "situation", "context"),
        "what_happened": find_col(raw_headers, "what happened", "event", "description"),
        "impact": find_col(raw_headers, "impact", "consequence"),
        "root_cause": find_col(raw_headers, "root cause"),
        "recommendation": find_col(raw_headers, "recommendation", "action"),
        "keywords": find_col(raw_headers, "keyword", "tag"),
        "status": find_col(raw_headers, "status"),
    }

    def get(row, key):
        idx = col_map.get(key, -1)
        if idx < 0 or idx >= len(row):
            return ""
        return row[idx].strip()

    lessons = []
    for row in rows[1:]:
        context = get(row, "context")
        what_happened = get(row, "what_happened")
        if not context and not what_happened:
            continue

        description = context
        if what_happened and what_happened != context:
            description += f"\n\nWhat happened: {what_happened}" if description else what_happened

        title = context.split(".")[0].split("\n")[0].strip()[:100]
        if not title:
            title = what_happened.split(".")[0].strip()[:100]
        if not title:
            continue

        impact = get(row, "impact")
        lessons.append({
            "title": title,
            "description": description,
            "root_cause": get(row, "root_cause"),
            "recommendation": get(row, "recommendation"),
            "impact": impact,
            "work_type": "Pipeline Construction",
            "phase": get(row, "phase"),
            "discipline": map_discipline(get(row, "discipline") or get(row, "category")),
            "severity": map_severity(impact),
            "project": get(row, "project"),
            "location": get(row, "region"),
            "keywords": get(row, "keywords"),
        })

    return lessons, None
