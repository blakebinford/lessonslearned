import io
import logging
from django.db.models import Count, Q
from django.utils import timezone
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from .models import Organization, Lesson, SOWAnalysis
from .serializers import (
    OrganizationSerializer,
    LessonSerializer,
    SOWAnalyzeRequestSerializer,
    ChatMessageSerializer,
)
from .parsers import parse_xlsx, parse_csv
from . import ai

logger = logging.getLogger(__name__)


class OrganizationViewSet(viewsets.ModelViewSet):
    serializer_class = OrganizationSerializer

    def get_queryset(self):
        return Organization.objects.filter(created_by=self.request.user)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class LessonViewSet(viewsets.ModelViewSet):
    serializer_class = LessonSerializer

    def get_queryset(self):
        qs = Lesson.objects.filter(organization__created_by=self.request.user)
        org_id = self.request.query_params.get("org")
        if org_id:
            qs = qs.filter(organization_id=org_id)
        # Filters
        discipline = self.request.query_params.get("discipline")
        if discipline:
            qs = qs.filter(discipline=discipline)
        severity = self.request.query_params.get("severity")
        if severity:
            qs = qs.filter(severity=severity)
        work_type = self.request.query_params.get("work_type")
        if work_type:
            qs = qs.filter(work_type=work_type)
        search = self.request.query_params.get("search")
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(title__icontains=search)
                | Q(description__icontains=search)
                | Q(root_cause__icontains=search)
                | Q(recommendation__icontains=search)
                | Q(keywords__icontains=search)
                | Q(project__icontains=search)
                | Q(location__icontains=search)
            )
        return qs

    def perform_create(self, serializer):
        org_id = self.request.data.get("organization")
        org = get_object_or_404(Organization, id=org_id, created_by=self.request.user)
        serializer.save(organization=org, created_by=self.request.user)

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        """Delete multiple lessons by ID. All must belong to the user's org."""
        ids = request.data.get("ids", [])
        if not ids or not isinstance(ids, list):
            return Response({"error": "ids list is required"}, status=400)
        qs = Lesson.objects.filter(
            id__in=ids, organization__created_by=request.user
        )
        found = set(qs.values_list("id", flat=True))
        missing = set(ids) - found
        if missing:
            return Response(
                {"error": f"Lessons not found or not authorized: {sorted(missing)}"},
                status=403,
            )
        count = qs.count()
        qs.delete()
        return Response({"deleted": count})

    @action(detail=False, methods=["post"], url_path="bulk-update")
    def bulk_update(self, request):
        """Update fields on multiple lessons. All must belong to the user's org."""
        ids = request.data.get("ids", [])
        fields = request.data.get("fields", {})
        if not ids or not isinstance(ids, list):
            return Response({"error": "ids list is required"}, status=400)
        if not fields or not isinstance(fields, dict):
            return Response({"error": "fields dict is required"}, status=400)
        allowed = {"severity", "work_type", "discipline", "phase", "environment"}
        invalid = set(fields.keys()) - allowed
        if invalid:
            return Response(
                {"error": f"Cannot bulk-update fields: {sorted(invalid)}"},
                status=400,
            )
        qs = Lesson.objects.filter(
            id__in=ids, organization__created_by=request.user
        )
        found = set(qs.values_list("id", flat=True))
        missing = set(ids) - found
        if missing:
            return Response(
                {"error": f"Lessons not found or not authorized: {sorted(missing)}"},
                status=403,
            )
        count = qs.update(**fields)
        return Response({"updated": count})

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_file(self, request):
        """Import lessons from an XLSX or CSV file."""
        org_id = request.data.get("organization")
        if not org_id:
            return Response({"error": "organization is required"}, status=400)
        org = get_object_or_404(Organization, id=org_id, created_by=request.user)

        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"error": "No file uploaded"}, status=400)

        file_bytes = uploaded.read()
        ext = uploaded.name.rsplit(".", 1)[-1].lower()

        if ext in ("xlsx", "xls", "xlsm"):
            parsed, error = parse_xlsx(file_bytes)
        elif ext in ("csv", "tsv"):
            parsed, error = parse_csv(file_bytes)
        else:
            return Response({"error": f"Unsupported file type: .{ext}"}, status=400)

        if error:
            return Response({"error": error}, status=400)
        if not parsed:
            return Response({"error": "No valid lessons found in file"}, status=400)

        # Create lesson objects
        created = 0
        for lesson_data in parsed:
            Lesson.objects.create(
                organization=org,
                created_by=request.user,
                **lesson_data,
            )
            created += 1

        return Response({
            "imported": created,
            "total_in_file": len(parsed),
            "filename": uploaded.name,
        })

    @action(detail=False, methods=["get"])
    def stats(self, request):
        """Lightweight aggregation stats for the lessons dashboard."""
        org_id = request.query_params.get("org")
        qs = Lesson.objects.filter(organization__created_by=request.user)
        if org_id:
            qs = qs.filter(organization_id=org_id)

        total = qs.count()

        by_severity = {}
        for row in qs.values("severity").annotate(n=Count("id")):
            by_severity[row["severity"]] = row["n"]

        by_discipline = {}
        for row in (
            qs.exclude(discipline="")
            .values("discipline")
            .annotate(n=Count("id"))
            .order_by("-n")[:3]
        ):
            by_discipline[row["discipline"]] = row["n"]

        by_work_type = {}
        for row in (
            qs.exclude(work_type="")
            .values("work_type")
            .annotate(n=Count("id"))
            .order_by("-n")[:3]
        ):
            by_work_type[row["work_type"]] = row["n"]

        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if now.month == 1:
            last_month_start = month_start.replace(year=now.year - 1, month=12)
        else:
            last_month_start = month_start.replace(month=now.month - 1)

        this_month = qs.filter(created_at__gte=month_start).count()
        last_month = qs.filter(
            created_at__gte=last_month_start, created_at__lt=month_start
        ).count()

        return Response({
            "total": total,
            "by_severity": by_severity,
            "by_discipline": by_discipline,
            "by_work_type": by_work_type,
            "this_month": this_month,
            "last_month": last_month,
        })


class SOWAnalysisViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only viewset for past SOW analyses."""

    def get_queryset(self):
        qs = SOWAnalysis.objects.filter(organization__created_by=self.request.user)
        org_id = self.request.query_params.get("org")
        if org_id:
            qs = qs.filter(organization_id=org_id)
        return qs

    def get_serializer_class(self):
        from .serializers import SOWAnalysisSerializer
        return SOWAnalysisSerializer


@api_view(["POST"])
def analyze_sow(request):
    """Run AI analysis on a scope of work against the lessons database."""
    try:
        print(">>> STEP 1: entered view")
        org_id = request.data.get("organization")
        sow_text = request.data.get("sow_text", "")
        work_type = request.data.get("work_type", "")
        filename = request.data.get("filename", "")
        print(f">>> STEP 2: org={org_id}, text_len={len(sow_text)}, wt={work_type}")

        if not org_id:
            return Response({"error": "organization is required"}, status=400)
        if not sow_text:
            return Response({"error": "sow_text is required"}, status=400)

        org = get_object_or_404(Organization, id=org_id, created_by=request.user)
        print(f">>> STEP 3: org found = {org.name}")

        lessons_qs = Lesson.objects.filter(organization=org).values(
            "id", "title", "description", "root_cause", "recommendation",
            "impact", "work_type", "phase", "discipline", "severity",
            "environment", "project", "location", "keywords",
        )
        lessons_list = list(lessons_qs)
        print(f">>> STEP 4: {len(lessons_list)} lessons loaded")

        org_profile = {"name": org.name, "profile_text": org.profile_text}
        print(">>> STEP 5: calling AI")

        results = ai.analyze_sow(sow_text, work_type, lessons_list, org_profile)
        print(f">>> STEP 6: AI returned, keys={list(results.keys()) if isinstance(results, dict) else 'not a dict'}")

        analysis = SOWAnalysis.objects.create(
            organization=org,
            filename=filename,
            sow_text=sow_text[:10000],
            work_type=work_type,
            results=results,
            created_by=request.user,
        )
        print(f">>> STEP 7: saved analysis id={analysis.id}")

        return Response({"id": analysis.id, "results": results})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


@api_view(["POST"])
def generate_deliverable(request):
    """Generate an on-demand deliverable from an existing SOW analysis."""
    analysis_id = request.data.get("analysis_id")
    deliverable_type = request.data.get("deliverable_type", "")
    params = request.data.get("params", {})

    if not analysis_id:
        return Response({"error": "analysis_id is required"}, status=400)
    if deliverable_type not in ai.DELIVERABLE_TYPES:
        return Response(
            {"error": f"Invalid deliverable_type. Must be one of: {sorted(ai.DELIVERABLE_TYPES)}"},
            status=400,
        )

    analysis = get_object_or_404(
        SOWAnalysis, id=analysis_id, organization__created_by=request.user
    )

    # Load lessons for the organization
    lessons_list = list(
        Lesson.objects.filter(organization=analysis.organization).values(
            "id", "title", "description", "root_cause", "recommendation",
            "impact", "work_type", "phase", "discipline", "severity",
            "environment", "project", "location", "keywords",
        )
    )

    org_profile = {
        "name": analysis.organization.name,
        "profile_text": analysis.organization.profile_text,
    }

    try:
        context = ai._build_analysis_context(analysis, lessons_list, org_profile)
        content = ai.generate_deliverable(deliverable_type, context, params)
    except Exception as e:
        logger.exception("Deliverable generation failed")
        return Response({"error": str(e)}, status=500)

    # Persist the deliverable in the analysis results JSON
    results = analysis.results or {}
    if "deliverables" not in results:
        results["deliverables"] = {}
    results["deliverables"][deliverable_type] = content
    analysis.results = results
    analysis.save(update_fields=["results"])

    return Response({
        "deliverable_type": deliverable_type,
        "content": content,
    })


@api_view(["POST"])
def upload_sow_file(request):
    """Extract text from an uploaded SOW document."""
    uploaded = request.FILES.get("file")
    if not uploaded:
        return Response({"error": "No file uploaded"}, status=400)

    ext = uploaded.name.rsplit(".", 1)[-1].lower()
    file_bytes = uploaded.read()

    text = ""
    if ext == "txt":
        text = file_bytes.decode("utf-8-sig")
    elif ext == "docx":
        try:
            import docx
            import io
            doc = docx.Document(io.BytesIO(file_bytes))
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            return Response({"error": "python-docx not installed on server"}, status=500)
    elif ext == "pdf":
        try:
            import pymupdf
            doc = pymupdf.open(stream=file_bytes, filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
        except ImportError:
            return Response(
                {"error": "pymupdf not installed. Install with: pip install pymupdf"},
                status=500,
            )
    else:
        return Response({"error": f"Unsupported file type: .{ext}"}, status=400)

    return Response({
        "text": text,
        "filename": uploaded.name,
        "length": len(text),
    })


@api_view(["POST"])
def export_sow_xlsx(request):
    """Export an SOW analysis as a styled Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    analysis_id = request.data.get("analysis_id")
    if not analysis_id:
        return Response({"error": "analysis_id is required"}, status=400)

    analysis = get_object_or_404(
        SOWAnalysis, id=analysis_id, organization__created_by=request.user
    )
    results = analysis.results or {}
    matches = results.get("matches", [])
    recommendations = results.get("recommendations", [])
    gaps = results.get("gaps", [])

    # Look up lesson details for each match
    lesson_ids = [m.get("lessonId") for m in matches if m.get("lessonId")]
    lessons_by_id = {
        l.id: l
        for l in Lesson.objects.filter(
            id__in=lesson_ids, organization=analysis.organization
        )
    }

    wb = Workbook()

    # --- Styles ---
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # ──────────────────────────────────────────────
    # Sheet 1: Applicable Lessons
    # ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Applicable Lessons"
    columns = [
        "Relevance", "Lesson Title", "Discipline", "Work Type",
        "Project", "Why It Applies", "Recommendation",
    ]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    widths = [12, 30, 16, 16, 20, 45, 45]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w

    row = 2
    for level_idx, level in enumerate(["High", "Medium", "Low"]):
        group = [m for m in matches if m.get("relevance") == level]
        if not group:
            continue
        # Blank separator row between groups (skip before first group)
        if level_idx > 0 and row > 2:
            row += 1
        for m in group:
            lesson = lessons_by_id.get(m.get("lessonId"))
            ws1.cell(row=row, column=1, value=level).font = cell_font
            ws1.cell(row=row, column=1).alignment = cell_alignment
            ws1.cell(row=row, column=2, value=lesson.title if lesson else str(m.get("lessonId", ""))).font = cell_font
            ws1.cell(row=row, column=2).alignment = cell_alignment
            ws1.cell(row=row, column=3, value=lesson.discipline if lesson else "").font = cell_font
            ws1.cell(row=row, column=3).alignment = cell_alignment
            ws1.cell(row=row, column=4, value=lesson.work_type if lesson else "").font = cell_font
            ws1.cell(row=row, column=4).alignment = cell_alignment
            ws1.cell(row=row, column=5, value=lesson.project if lesson else "").font = cell_font
            ws1.cell(row=row, column=5).alignment = cell_alignment
            ws1.cell(row=row, column=6, value=m.get("reason", "")).font = cell_font
            ws1.cell(row=row, column=6).alignment = cell_alignment
            ws1.cell(row=row, column=7, value=lesson.recommendation if lesson else "").font = cell_font
            ws1.cell(row=row, column=7).alignment = cell_alignment
            # Apply border to all cells in the row
            for c in range(1, 8):
                ws1.cell(row=row, column=c).border = thin_border
            row += 1

    # ──────────────────────────────────────────────
    # Sheet 2: Recommendations
    # ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Recommendations")
    cell = ws2.cell(row=1, column=1, value="Recommendations")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    ws2.column_dimensions["A"].width = 80
    for i, rec in enumerate(recommendations, 1):
        ws2.cell(row=i + 1, column=1, value=f"{i}. {rec}").font = cell_font
        ws2.cell(row=i + 1, column=1).alignment = cell_alignment

    # ──────────────────────────────────────────────
    # Sheet 3: Gaps
    # ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Gaps")
    cell = ws3.cell(row=1, column=1, value="Gaps")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    ws3.column_dimensions["A"].width = 80
    for i, gap in enumerate(gaps, 1):
        ws3.cell(row=i + 1, column=1, value=gap).font = cell_font
        ws3.cell(row=i + 1, column=1).alignment = cell_alignment

    # Write to buffer and return
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = analysis.filename or "sow-analysis"
    if filename.rsplit(".", 1)[-1].lower() in ("docx", "pdf", "txt", "doc"):
        filename = filename.rsplit(".", 1)[0]
    safe_filename = filename.replace('"', "")

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{safe_filename} - SOW Analysis.xlsx"'
    return response


@api_view(["POST"])
def chat_analyst(request):
    """AI chat analyst endpoint."""
    serializer = ChatMessageSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    org_id = request.data.get("organization")
    if not org_id:
        return Response({"error": "organization is required"}, status=400)
    org = get_object_or_404(Organization, id=org_id, created_by=request.user)

    message = serializer.validated_data["message"]
    history = serializer.validated_data.get("history", [])

    lessons = list(
        Lesson.objects.filter(organization=org).values(
            "id", "title", "description", "root_cause", "recommendation",
            "work_type", "discipline", "severity", "environment",
        )
    )

    org_profile = {"name": org.name, "profile_text": org.profile_text}

    try:
        response_text = ai.chat_with_analyst(message, history, lessons, org_profile)
    except Exception as e:
        logger.exception("Chat analyst failed")
        return Response({"error": str(e)}, status=500)

    return Response({"response": response_text})


@api_view(["POST"])
@permission_classes([permissions.AllowAny])
def register(request):
    """Simple user registration endpoint."""
    from django.contrib.auth.models import User

    username = request.data.get("username", "").strip()
    password = request.data.get("password", "")
    email = request.data.get("email", "").strip()

    if not username or not password:
        return Response({"error": "username and password required"}, status=400)
    if User.objects.filter(username=username).exists():
        return Response({"error": "Username already taken"}, status=400)

    user = User.objects.create_user(username=username, password=password, email=email)

    # Create default org
    Organization.objects.create(name=f"{username}'s Organization", created_by=user)

    # Create auth token
    from rest_framework.authtoken.models import Token
    token, _ = Token.objects.get_or_create(user=user)

    return Response({"token": token.key, "user_id": user.id, "username": user.username})


@api_view(["POST"])
@permission_classes([permissions.AllowAny])
def login_view(request):
    """Token-based login."""
    from django.contrib.auth import authenticate
    from rest_framework.authtoken.models import Token

    username = request.data.get("username", "")
    password = request.data.get("password", "")

    user = authenticate(username=username, password=password)
    if not user:
        return Response({"error": "Invalid credentials"}, status=401)

    token, _ = Token.objects.get_or_create(user=user)
    return Response({"token": token.key, "user_id": user.id, "username": user.username})